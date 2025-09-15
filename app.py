# from flask import Flask, render_template, request, send_file, url_for
# import pyodbc
# import pandas as pd
# import os
# from datetime import datetime, timedelta
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill
# import shutil
# from tag_map import tag_map  # ✅ DB column → Excel column map
# from apscheduler.schedulers.background import BackgroundScheduler

# app = Flask(__name__)

# # ---------------- Database Configuration ----------------
# DB_SERVER = r"DESKTOP-KTH18G4\SQLEXPRESS"   # Server name
# DB_NAME = "ITPL"                            # Database name
# TABLE_NAME = "PLC_Report1"                  # Table name

# # Connection string (SQL Authentication)
# conn_str = (
#     "DRIVER={ODBC Driver 17 for SQL Server};"
#     f"SERVER={DB_SERVER};"
#     f"DATABASE={DB_NAME};"
#     "UID=sa;"
#     "PWD=Admin@123#;"
# )

# # ---------------- File Paths ----------------
# TEMPLATE_PATH = r"D:\Summerpur_Reports\Summerpur_Reports\Unit Reco Count Template.xlsx"
# REPORTS_DIR = r"D:\Summerpur_Reports\Summerpur_Reports\Reports"

# os.makedirs(REPORTS_DIR, exist_ok=True)

# # ---------------- Report Generation ----------------
# def generate_report(base_date, selected_shift):
#     """Generate report for given date & shift"""

#     # Shift time ranges
#     if selected_shift == "Shift-A":
#         start_time = base_date.replace(hour=7, minute=0, second=0)
#         end_time = base_date.replace(hour=14, minute=59, second=59)
#     elif selected_shift == "Shift-B":
#         start_time = base_date.replace(hour=15, minute=0, second=0)
#         end_time = base_date.replace(hour=22, minute=59, second=59)
#     elif selected_shift == "Shift-C":
#         start_time = base_date.replace(hour=23, minute=0, second=0)
#         end_time = (base_date + timedelta(days=1)).replace(hour=6, minute=59, second=59)
#     else:  # ✅ Full-Day (07:05 → Next Day 07:00)
#         start_time = base_date.replace(hour=7, minute=5, second=0)
#         end_time = (base_date + timedelta(days=1)).replace(hour=7, minute=0, second=0)

#     # SQL Query
#     tag_columns = list(tag_map.keys())     
#     query = f"""
#         SELECT DT, {", ".join(tag_columns)}
#         FROM {TABLE_NAME}
#         WHERE DT BETWEEN ? AND ?
#         ORDER BY DT
#     """

#     try:
#         with pyodbc.connect(conn_str) as conn:
#             df = pd.read_sql(query, conn, params=[start_time, end_time])
#     except Exception as e:
#         print(f"❌ Database error: {e}")
#         return None

#     if df.empty:
#         print(f"❌ No data found for {selected_shift} on {base_date.date()}")
#         return None

#     # Resample to 5-min interval & round
#     df["DT"] = pd.to_datetime(df["DT"])
#     df = df.set_index("DT").resample("5T").mean().reset_index()
#     df = df.round(0).astype(int, errors="ignore")

#     # File save
#     date_str = base_date.strftime("%d-%m-%Y")
#     output_file = os.path.join(REPORTS_DIR, f"Daily_Report-{date_str}_{selected_shift}.xlsx")
#     shutil.copy(TEMPLATE_PATH, output_file)

#     # Write into Excel
#     wb = load_workbook(output_file)
#     ws = wb.active

#     row = 2
#     for _, r in df.iterrows():
#         ws[f"B{row}"] = pd.to_datetime(r["DT"]).strftime("%d-%m-%Y %H:%M")
#         ws[f"C{row}"] = selected_shift
#         for tag, excel_col in tag_map.items():
#             if tag in r:
#                 ws[f"{excel_col}{row}"] = r[tag]
#         row += 1

#     # ✅ TOTAL row banayein (last row ka data copy karke)
#     total_row = row
#     ws[f"B{total_row}"] = "TOTAL"
#     ws[f"C{total_row}"] = selected_shift

#     last_row_data = df.iloc[-1]  # ✅ last row ka data

#     for tag, excel_col in tag_map.items():
#         if tag in last_row_data:
#             ws[f"{excel_col}{total_row}"] = last_row_data[tag]

#     # Style TOTAL row
#     bold_font = Font(bold=True)
#     green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
#     for col in range(2, ws.max_column + 1):
#         cell = ws.cell(row=total_row, column=col)
#         cell.font = bold_font
#         cell.fill = green_fill

#     wb.save(output_file)
#     print(f"✅ Report generated: {output_file}")
#     return output_file


# # ---------------- Flask Routes ----------------
# @app.route("/", methods=["GET", "POST"])
# def index():
#     if request.method == "POST":
#         selected_date = request.form["date"]
#         selected_shift = request.form["shift"]

#         try:
#             base_date = datetime.strptime(selected_date, "%Y-%m-%d")
#             output_file = generate_report(base_date, selected_shift)

#             if not output_file:
#                 return render_template("index.html", message="❌ No data found.")

#             download_link = url_for("download_report", filename=os.path.basename(output_file))
#             return render_template("index.html", message="✅ Report generated!", download_link=download_link)

#         except Exception as e:
#             return render_template("index.html", message=f"❌ Error: {str(e)}")

#     return render_template("index.html")


# @app.route("/download/<filename>")
# def download_report(filename):
#     file_path = os.path.join(REPORTS_DIR, filename)
#     return send_file(file_path, as_attachment=True)


# # ---------------- Scheduler ----------------
# def schedule_daily_reports():
#     """Auto-generate reports for all shifts + Full-Day at 7 AM"""
#     yesterday = datetime.now() - timedelta(days=1)
#     for shift in ["Shift-A", "Shift-B", "Shift-C", "Full-Day"]:
#         generate_report(yesterday, shift)

# scheduler = BackgroundScheduler()
# scheduler.add_job(schedule_daily_reports, "cron", hour=7, minute=0)
# scheduler.start()


# if __name__ == "__main__":
#     app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)

from flask import Flask, render_template, request, send_file, url_for
import pyodbc
import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import shutil
from tag_map import tag_map  # ✅ DB column → Excel column map
from apscheduler.schedulers.background import BackgroundScheduler

app = Flask(__name__)

# ---------------- Database Configuration ----------------
DB_SERVER = r"DESKTOP-KTH18G4\SQLEXPRESS"   # Server name
DB_NAME = "ITPL"                            # Database name
TABLE_NAME = "PLC_Report1"                  # Table name

# Connection string (SQL Authentication)
conn_str = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    f"SERVER={DB_SERVER};"
    f"DATABASE={DB_NAME};"
    "UID=sa;"
    "PWD=Admin@123#;"
)

# ---------------- File Paths ----------------
TEMPLATE_PATH = r"D:\Summerpur_Reports\Summerpur_Reports\Unit Reco Count Template.xlsx"
REPORTS_DIR = r"D:\Summerpur_Reports\Summerpur_Reports\Reports"

os.makedirs(REPORTS_DIR, exist_ok=True)

# ---------------- Report Generation ----------------
def generate_report(base_date, selected_shift):
    """Generate report for given date & shift"""

    # Shift time ranges
    if selected_shift == "Shift-A":
        start_time = base_date.replace(hour=7, minute=0, second=0)
        end_time = base_date.replace(hour=14, minute=59, second=59)
    elif selected_shift == "Shift-B":
        start_time = base_date.replace(hour=15, minute=0, second=0)
        end_time = base_date.replace(hour=22, minute=59, second=59)
    elif selected_shift == "Shift-C":
        start_time = base_date.replace(hour=23, minute=0, second=0)
        end_time = (base_date + timedelta(days=1)).replace(hour=6, minute=59, second=59)
    else:  # ✅ Full-Day (07:05 → Next Day 07:00)
        start_time = base_date.replace(hour=7, minute=5, second=0)
        end_time = (base_date + timedelta(days=1)).replace(hour=7, minute=0, second=0)

    # SQL Query
    tag_columns = list(tag_map.keys())     
    query = f"""
        SELECT DT, {", ".join(tag_columns)}
        FROM {TABLE_NAME}
        WHERE DT BETWEEN ? AND ?
        ORDER BY DT
    """

    try:
        with pyodbc.connect(conn_str) as conn:
            df = pd.read_sql(query, conn, params=[start_time, end_time])
    except Exception as e:
        print(f"❌ Database error: {e}")
        return None

    if df.empty:
        print(f"❌ No data found for {selected_shift} on {base_date.date()}")
        return None

    # Resample to 5-min interval & round
    df["DT"] = pd.to_datetime(df["DT"])
    df = df.set_index("DT").resample("5T").mean().reset_index()
    df = df.round(0).astype(int, errors="ignore")

    # File save
    date_str = base_date.strftime("%d-%m-%Y")
    output_file = os.path.join(REPORTS_DIR, f"Daily_Report-{date_str}_{selected_shift}.xlsx")
    shutil.copy(TEMPLATE_PATH, output_file)

    # Write into Excel
    wb = load_workbook(output_file)
    ws = wb.active

    row = 2
    for _, r in df.iterrows():
        ws[f"B{row}"] = pd.to_datetime(r["DT"]).strftime("%d-%m-%Y %H:%M")
        ws[f"C{row}"] = selected_shift
        for tag, excel_col in tag_map.items():
            if tag in r:
                ws[f"{excel_col}{row}"] = r[tag]
        row += 1

    # ✅ TOTAL row banayein (last row ka data copy karke)
    total_row = row
    ws[f"B{total_row}"] = "TOTAL"
    ws[f"C{total_row}"] = selected_shift

    last_row_data = df.iloc[-1]  # ✅ last row ka data

    for tag, excel_col in tag_map.items():
        if tag in last_row_data:
            ws[f"{excel_col}{total_row}"] = last_row_data[tag]

    # Style TOTAL row
    bold_font = Font(bold=True)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = bold_font
        cell.fill = green_fill

    wb.save(output_file)
    print(f"✅ Report generated: {output_file}")
    return output_file


# ---------------- Flask Routes ----------------
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        selected_date = request.form["date"]
        selected_shift = request.form["shift"]

        try:
            base_date = datetime.strptime(selected_date, "%Y-%m-%d")
            output_file = generate_report(base_date, selected_shift)

            if not output_file:
                return render_template("index.html", message="❌ No data found.")

            download_link = url_for("download_report", filename=os.path.basename(output_file))
            return render_template("index.html", message="✅ Report generated!", download_link=download_link)

        except Exception as e:
            return render_template("index.html", message=f"❌ Error: {str(e)}")

    return render_template("index.html")


@app.route("/download/<filename>")
def download_report(filename):
    file_path = os.path.join(REPORTS_DIR, filename)
    return send_file(file_path, as_attachment=True)


# ---------------- Scheduler ----------------
def schedule_daily_reports():
    """Auto-generate reports for all shifts + Full-Day at 7 AM"""
    yesterday = datetime.now() - timedelta(days=1)
    for shift in ["Shift-A", "Shift-B", "Shift-C", "Full-Day"]:
        generate_report(yesterday, shift)

scheduler = BackgroundScheduler()
scheduler.add_job(schedule_daily_reports, "cron", hour=7, minute=0)
scheduler.start()

# ---------------- Run Flask ----------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)
